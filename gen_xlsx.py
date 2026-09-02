import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

filepath = r"D:\web\StackHK13\AI_X_关注清单.xlsx"

wb = openpyxl.Workbook()

# Colors
HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color="0F172A")
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=9, color="64748B")
CAT_FILL_MAP = {
    "A.大模型基座": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "B.AI工具": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "C.AI SaaS": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
    "D.大模型中转站/API网关": PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid"),
    "E.AI媒体": PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid"),
    "F.AI评测/榜单": PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"),
    "G.AI头部自媒体/KOL": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
}
CAT_FONT = Font(name="Microsoft YaHei", size=9, bold=True, color="1E293B")
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

# Data
data = [
    # A.大模型基座 50
    ["A.大模型基座", "1", "OpenAI", "@OpenAI", "GPT-5 / GPT-4o / o3", "美国 旧金山", "EN", "https://x.com/OpenAI", "必关第一位，所有新模型/发布会首发，适合转载+蹭热点"],
    ["A.大模型基座", "2", "Anthropic", "@AnthropicAI", "Claude 4.6 Opus/Sonnet", "美国 旧金山", "EN", "https://x.com/AnthropicAI", "安全+代码能力最强，Claude Code必跟，适合深度内容"],
    ["A.大模型基座", "3", "Anthropic Claude", "@claudeai", "Claude官方助手号", "美国", "EN", "https://x.com/claudeai", "产品更新/use case"],
    ["A.大模型基座", "4", "Google DeepMind", "@GoogleDeepMind", "Gemini 3.1 / Gemma", "美国/英国", "EN", "https://x.com/GoogleDeepMind", "论文+多模态里程碑，适合学术转载"],
    ["A.大模型基座", "5", "Google AI", "@GoogleAI", "Gemini / Search AI", "美国", "EN", "https://x.com/GoogleAI", "Google全产品AI整合"],
    ["A.大模型基座", "6", "Meta AI", "@AIatMeta", "Llama 4 Maverick", "美国", "EN", "https://x.com/AIatMeta", "开源权重领袖，开源圈必关"],
    ["A.大模型基座", "7", "xAI", "@xai", "Grok 4 / Grok 5", "美国", "EN", "https://x.com/xai", "马斯克系，实时推理+热点强"],
    ["A.大模型基座", "8", "Mistral AI", "@MistralAI", "Mistral Large 3 / Mixtral", "法国 巴黎", "EN/FR", "https://x.com/MistralAI", "欧洲之光，开源+主权AI"],
    ["A.大模型基座", "9", "DeepSeek", "@deepseek_ai", "DeepSeek-R2 / V3", "中国 杭州", "EN/ZH", "https://x.com/deepseek_ai", "性价比推理之王，成本向必关"],
    ["A.大模型基座", "10", "Cohere", "@cohere", "Command R+", "加拿大", "EN", "https://x.com/cohere", "企业RAG/搜索专用"],
    ["A.大模型基座", "11", "Alibaba Qwen", "@Alibaba_Qwen", "Qwen3", "中国 杭州", "EN/ZH", "https://x.com/Alibaba_Qwen", "阿里通义千问，开源主力"],
    ["A.大模型基座", "12", "ByteDance Seed", "@ByteDanceResearch", "Doubao / Seed", "中国 北京", "EN", "https://x.com/ByteDanceResearch", "豆包，字节研究号"],
    ["A.大模型基座", "13", "Baidu 百度", "@Baidu_Inc", "ERNIE 5.0 文心", "中国 北京", "EN/ZH", "https://x.com/Baidu_Inc", "集团号为主，模型信息少"],
    ["A.大模型基座", "14", "Tencent Hunyuan", "@TencentAI", "混元 Hunyuan", "中国 深圳", "EN", "https://x.com/TencentAI", "腾讯混元"],
    ["A.大模型基座", "15", "Zhipu AI 智谱", "@ZhipuAI", "GLM-4.5", "中国 北京", "EN/ZH", "https://x.com/ZhipuAI", "GLM，清华系"],
    ["A.大模型基座", "16", "Moonshot 月之暗面", "@Kimi_Moonshot", "Kimi K2", "中国 北京", "EN/ZH", "https://x.com/Kimi_Moonshot", "长上下文代表"],
    ["A.大模型基座", "17", "MiniMax", "@MiniMaxAI_Official", "MiniMax M1 / Video", "中国 上海", "EN", "https://x.com/MiniMaxAI_Official", "多模态+视频"],
    ["A.大模型基座", "18", "01.AI 零一万物", "@01AI_Yi", "Yi-Large", "中国 北京", "EN", "https://x.com/01AI_Yi", "李开复团队"],
    ["A.大模型基座", "19", "Baichuan 百川", "@BaichuanAI", "Baichuan 4", "中国 北京", "EN", "https://x.com/BaichuanAI", "王小川团队"],
    ["A.大模型基座", "20", "StepFun 阶跃星辰", "@StepFunAI", "Step 2", "中国 上海", "EN", "https://x.com/StepFunAI", "上海阶跃"],
    ["A.大模型基座", "21", "AI21 Labs", "@AI21Labs", "Jamba", "以色列/美国", "EN", "https://x.com/AI21Labs", "企业LLM"],
    ["A.大模型基座", "22", "Aleph Alpha", "@aleph__alpha", "Luminous", "德国", "EN/DE", "https://x.com/aleph__alpha", "德国主权AI"],
    ["A.大模型基座", "23", "Stability AI", "@StabilityAI", "Stable LM / SD3", "英国", "EN", "https://x.com/StabilityAI", "开源生图+LLM"],
    ["A.大模型基座", "24", "Inflection AI", "@inflectionAI", "Pi", "美国", "EN", "https://x.com/inflectionAI", "情感陪伴AI"],
    ["A.大模型基座", "25", "Character.AI", "@character_ai", "Character", "美国", "EN", "https://x.com/character_ai", "角色扮演C端"],
    ["A.大模型基座", "26", "Safe Superintelligence", "@SSI", "SSI Lab", "美国", "EN", "https://x.com/SSI", "Ilya新公司，安全超级智能"],
    ["A.大模型基座", "27", "Thinking Machines", "@thinkymachines", "Mira Murati团队", "美国", "EN", "https://x.com/thinkymachines", "前OpenAI CTO新 lab"],
    ["A.大模型基座", "28", "Reflection AI", "@reflectionAI_", "Reflection 70B", "美国", "EN", "https://x.com/reflectionAI_", "开源挑战者"],
    ["A.大模型基座", "29", "World Labs", "@theworldlabs", "Spatial AI", "美国", "EN", "https://x.com/theworldlabs", "李飞飞空间智能"],
    ["A.大模型基座", "30", "Hugging Face", "@huggingface", "Hub/平台", "美国/法国", "EN", "https://x.com/huggingface", "开源模型GitHub，必关"],
    ["A.大模型基座", "31", "Microsoft AI", "@MicrosoftAI", "Copilot / Phi-4", "美国", "EN", "https://x.com/MicrosoftAI", "企业分发最强"],
    ["A.大模型基座", "32", "Amazon AGI/Bedrock", "@awscloud", "Nova / Bedrock", "美国", "EN", "https://x.com/awscloud", "AWS AI平台"],
    ["A.大模型基座", "33", "IBM watsonx", "@IBM", "Granite", "美国", "EN", "https://x.com/IBM", "企业AI"],
    ["A.大模型基座", "34", "NVIDIA", "@nvidia", "Nemotron", "美国", "EN", "https://x.com/nvidia", "算力霸主"],
    ["A.大模型基座", "35", "Apple Intelligence", "@Apple", "Apple FM", "美国", "EN", "https://x.com/Apple", "端侧AI"],
    ["A.大模型基座", "36", "Databricks Mosaic", "@databricks", "DBRX", "美国", "EN", "https://x.com/databricks", "数据+AI湖仓"],
    ["A.大模型基座", "37", "Upstage", "@UpstageAI", "Solar", "韩国", "EN/KR", "https://x.com/UpstageAI", "韩国头号LLM"],
    ["A.大模型基座", "38", "LG AI Research", "@LG_AI_Research", "EXAONE", "韩国", "EN/KR", "https://x.com/LG_AI_Research", "韩国LG"],
    ["A.大模型基座", "39", "Cohere (复核)", "@cohere", "-", "加拿大", "EN", "https://x.com/cohere", "见上"],
    ["A.大模型基座", "40", "Perplexity 归入工具但属模型", "@perplexity_ai", "Sonar", "美国", "EN", "https://x.com/perplexity_ai", "AI搜索模型"],

    # B. AI工具
    ["B.AI工具", "1", "Cursor", "@cursor_ai", "AI代码编辑器", "美国", "EN", "https://x.com/cursor_ai", "目前最火AI编程工具，必关"],
    ["B.AI工具", "2", "GitHub Copilot", "@github", "Copilot", "美国", "EN", "https://x.com/github", "编程工具官方"],
    ["B.AI工具", "3", "Cognition Devin", "@cognition_labs", "Devin Agent", "美国", "EN", "https://x.com/cognition_labs", "自主编程智能体"],
    ["B.AI工具", "4", "Replit", "@Replit", "Replit Agent", "美国", "EN", "https://x.com/Replit", "云端AI编程"],
    ["B.AI工具", "5", "Lovable", "@lovable_dev", "AI建站", "瑞典", "EN", "https://x.com/lovable_dev", "自然语言建站爆款"],
    ["B.AI工具", "6", "Bolt.new", "@boltdotnew", "AI建站", "美国", "EN", "https://x.com/boltdotnew", "StackBlitz出品"],
    ["B.AI工具", "7", "v0 by Vercel", "@v0", "AI前端生成", "美国", "EN", "https://x.com/v0", "前端AI生成"],
    ["B.AI工具", "8", "Notion AI", "@NotionHQ", "AI笔记", "美国", "EN", "https://x.com/NotionHQ", "AI效率工具代表"],
    ["B.AI工具", "9", "Canva Magic", "@canva", "AI设计", "澳洲", "EN", "https://x.com/canva", "AI设计SaaS"],
    ["B.AI工具", "10", "Figma AI", "@figma", "AI设计", "美国", "EN", "https://x.com/figma", "AI设计"],
    ["B.AI工具", "11", "Gamma", "@gammaapp", "AI PPT", "美国", "EN", "https://x.com/gammaapp", "AI演示文稿"],
    ["B.AI工具", "12", "Midjourney", "@midjourney", "V7 生图", "美国", "EN", "https://x.com/midjourney", "生图龙头"],
    ["B.AI工具", "13", "Runway", "@runwayml", "Gen-4 视频", "美国", "EN", "https://x.com/runwayml", "视频生成龙头"],
    ["B.AI工具", "14", "Pika", "@pika_labs", "Pika 2.0", "美国", "EN", "https://x.com/pika_labs", "视频生成"],
    ["B.AI工具", "15", "Luma AI", "@LumaLabsAI", "Dream Machine", "美国", "EN", "https://x.com/LumaLabsAI", "视频3D"],
    ["B.AI工具", "16", "Suno AI", "@suno_ai_", "Suno v4", "美国", "EN", "https://x.com/suno_ai_", "音乐生成"],
    ["B.AI工具", "17", "Udio", "@udiomusic", "Udio", "美国", "EN", "https://x.com/udiomusic", "音乐生成"],
    ["B.AI工具", "18", "ElevenLabs", "@elevenlabsio", "Eleven v3 语音", "英国/美国", "EN", "https://x.com/elevenlabsio", "语音合成第一"],
    ["B.AI工具", "19", "HeyGen", "@HeyGen_Official", "AI数字人", "美国", "EN", "https://x.com/HeyGen_Official", "数字人视频"],
    ["B.AI工具", "20", "Perplexity", "@perplexity_ai", "AI搜索", "美国", "EN", "https://x.com/perplexity_ai", "AI搜索，工具+模型"],
    ["B.AI工具", "21", "Genspark", "@GensparkAI", "AI Agent搜索", "美国", "EN", "https://x.com/GensparkAI", "新晋AI搜索"],
    ["B.AI工具", "22", "Grammarly", "@Grammarly", "AI写作", "美国", "EN", "https://x.com/Grammarly", "写作辅助"],
    ["B.AI工具", "23", "Jasper", "@jasperai", "AI营销写作", "美国", "EN", "https://x.com/jasperai", "营销SaaS"],

    # C. AI SaaS
    ["C.AI SaaS", "1", "Salesforce Einstein", "@salesforce", "AI CRM", "美国", "EN", "https://x.com/salesforce", "AI CRM龙头"],
    ["C.AI SaaS", "2", "HubSpot AI", "@HubSpot", "AI营销", "美国", "EN", "https://x.com/HubSpot", "营销SaaS"],
    ["C.AI SaaS", "3", "Intercom Fin", "@intercom", "AI客服", "美国", "EN", "https://x.com/intercom", "AI客服标杆"],
    ["C.AI SaaS", "4", "Zendesk AI", "@Zendesk", "AI客服", "美国", "EN", "https://x.com/Zendesk", "客服SaaS"],
    ["C.AI SaaS", "5", "Sierra", "@sierra", "AI客服智能体", "美国", "EN", "https://x.com/sierra", "Bret Taylor新公司"],
    ["C.AI SaaS", "6", "Decagon", "@decagonai", "AI客服", "美国", "EN", "https://x.com/decagonai", "AI客服新锐"],
    ["C.AI SaaS", "7", "Harvey", "@harvey__ai", "AI法律", "美国", "EN", "https://x.com/harvey__ai", "法律AI独角兽"],
    ["C.AI SaaS", "8", "Legora", "@LegoraHQ", "AI法律欧洲", "瑞典", "EN", "https://x.com/LegoraHQ", "欧洲法律AI"],
    ["C.AI SaaS", "9", "Glean", "@glean", "AI企业搜索", "美国", "EN", "https://x.com/glean", "企业知识搜索"],
    ["C.AI SaaS", "10", "Clay", "@clayhq", "AI销售拓客", "美国", "EN", "https://x.com/clayhq", "GTM神器"],
    ["C.AI SaaS", "11", "Scale AI", "@scale_AI", "数据标注/评测", "美国", "EN", "https://x.com/scale_AI", "数据层龙头"],
    ["C.AI SaaS", "12", "LangChain", "@LangChainAI", "Agent框架", "美国", "EN", "https://x.com/LangChainAI", "Agent开发必关"],
    ["C.AI SaaS", "13", "Weights & Biases", "@weights_biases", "MLOps", "美国", "EN", "https://x.com/weights_biases", "模型训练观测"],
    ["C.AI SaaS", "14", "Palantir", "@PalantirTech", "AI OS", "美国", "EN", "https://x.com/PalantirTech", "政府/企业AI"],

    # D. 中转站
    ["D.大模型中转站/API网关", "1", "OpenRouter", "@OpenRouterAI", "全球最大聚合API", "美国", "EN", "https://x.com/OpenRouterAI", "最全模型聚合，价格透明，必关"],
    ["D.大模型中转站/API网关", "2", "Portkey AI", "@PortkeyAI", "AI网关/路由", "美国/印度", "EN", "https://x.com/PortkeyAI", "网关+可观测+缓存"],
    ["D.大模型中转站/API网关", "3", "Helicone", "@helicone_ai", "LLM可观测网关", "美国", "EN", "https://x.com/helicone_ai", "开源可观测"],
    ["D.大模型中转站/API网关", "4", "LiteLLM / BerriAI", "@BerriAI", "统一API网关", "美国", "EN", "https://x.com/BerriAI", "开源统一接口"],
    ["D.大模型中转站/API网关", "5", "Cloudflare AI Gateway", "@cloudflare", "AI网关", "美国", "EN", "https://x.com/cloudflare", " Workers AI网关"],
    ["D.大模型中转站/API网关", "6", "Vercel AI Gateway", "@vercel", "AI SDK+网关", "美国", "EN", "https://x.com/vercel", "前端AI网关"],
    ["D.大模型中转站/API网关", "7", "Together AI", "@togethercompute", "推理云/聚合", "美国", "EN", "https://x.com/togethercompute", "开源模型推理便宜"],
    ["D.大模型中转站/API网关", "8", "Fireworks AI", "@FireworksAI_HQ", "推理云", "美国", "EN", "https://x.com/FireworksAI_HQ", "极速推理"],
    ["D.大模型中转站/API网关", "9", "Replicate", "@replicate", "模型API托管", "美国", "EN", "https://x.com/replicate", "一键API部署"],
    ["D.大模型中转站/API网关", "10", "Fal.ai", "@fal", "媒体模型API", "美国", "EN", "https://x.com/fal", "生图/视频API"],
    ["D.大模型中转站/API网关", "11", "GroqCloud", "@GroqInc", "极速推理", "美国", "EN", "https://x.com/GroqInc", "LPU推理最快"],
    ["D.大模型中转站/API网关", "12", "Poe by Quora", "@poe_platform", "模型聚合C端", "美国", "EN", "https://x.com/poe_platform", "多模型聊天聚合"],
    ["D.大模型中转站/API网关", "13", "SiliconFlow 硅基流动", "@SiliconFlowAI", "国内聚合", "中国", "EN/ZH", "https://x.com/SiliconFlowAI", "国内合规聚合，Qwen/DeepSeek免费"],
    ["D.大模型中转站/API网关", "14", "ModelScope 魔塔", "@ModelScope", "阿里模型聚合", "中国", "EN/ZH", "https://x.com/ModelScope", "阿里开源社区"],
    ["D.大模型中转站/API网关", "15", "Anyscale", "@anyscalecompute", "Ray/Endpoints", "美国", "EN", "https://x.com/anyscalecompute", "Ray分布式推理"],
    ["D.大模型中转站/API网关", "16", "Langfuse", "@langfuse", "LLM工程化", "德国", "EN", "https://x.com/langfuse", "开源LLM观测"],

    # E. 媒体
    ["E.AI媒体", "1", "MIT Tech Review", "@techreview", "科技深度", "美国", "EN", "https://x.com/techreview", "最权威AI解读"],
    ["E.AI媒体", "2", "The Verge", "@verge", "消费科技", "美国", "EN", "https://x.com/verge", "热点快"],
    ["E.AI媒体", "3", "WIRED", "@WIRED", "科技文化", "美国", "EN", "https://x.com/WIRED", "深度特写"],
    ["E.AI媒体", "4", "TechCrunch", "@TechCrunch", "创投融资", "美国", "EN", "https://x.com/TechCrunch", "融资首发"],
    ["E.AI媒体", "5", "VentureBeat", "@VentureBeat", "AI产业", "美国", "EN", "https://x.com/VentureBeat", "产业分析"],
    ["E.AI媒体", "6", "The Information", "@theinformation", "硅谷付费深度", "美国", "EN", "https://x.com/theinformation", "独家爆料多"],
    ["E.AI媒体", "7", "Ars Technica", "@arstechnica", "技术评测", "美国", "EN", "https://x.com/arstechnica", "技术细节"],
    ["E.AI媒体", "8", "SemiAnalysis", "@SemiAnalysis_", "芯片+基建研报", "美国", "EN", "https://x.com/SemiAnalysis_", "Dylan Patel 芯片一哥"],
    ["E.AI媒体", "9", "The Decoder", "@TheDecoderNews", "欧洲AI媒体", "德国", "EN", "https://x.com/TheDecoderNews", "欧洲视角"],
    ["E.AI媒体", "10", "Bloomberg Tech", "@technology", "财经科技", "美国", "EN", "https://x.com/technology", "财报/股价联动"],
    ["E.AI媒体", "11", "Testing Catalog", "@testingcatalog", "AI产品爆料", "全球", "EN", "https://x.com/testingcatalog", "新功能爆料最快"],

    # F. 评测
    ["F.AI评测/榜单", "1", "LM Arena / LMSYS", "@lmsysorg", "Chatbot Arena", "美国", "EN", "https://x.com/lmsysorg", "人类盲测金标准，必@"],
    ["F.AI评测/榜单", "2", "Artificial Analysis", "@ArtificialAnlys", "质量/价格/速度", "澳洲", "EN", "https://x.com/ArtificialAnlys", "最客观独立评测"],
    ["F.AI评测/榜单", "3", "Epoch AI", "@EpochAIResearch", "算力/趋势研报", "美国", "EN", "https://x.com/EpochAIResearch", "数据扎实"],
    ["F.AI评测/榜单", "4", "LiveBench", "@livebench_ai", "实时防污染榜", "美国", "EN", "https://x.com/livebench_ai", "防刷榜"],
    ["F.AI评测/榜单", "5", "Stanford HAI", "@StanfordHAI", "HELM学术评测", "美国", "EN", "https://x.com/StanfordHAI", "学术权威"],
    ["F.AI评测/榜单", "6", "Hugging Face", "@huggingface", "Open LLM LB", "美国", "EN", "https://x.com/huggingface", "开源榜单"],
    ["F.AI评测/榜单", "7", "Vellum AI", "@vellum_ai", "模型对比", "美国", "EN", "https://x.com/vellum_ai", "对比工具"],

    # G. KOL
    ["G.AI头部自媒体/KOL", "1", "Andrej Karpathy", "@karpathy", "AI第一博主", "美国", "EN", "https://x.com/karpathy", "必关，教程+洞察顶流"],
    ["G.AI头部自媒体/KOL", "2", "Yann LeCun", "@ylecun", "Meta 图灵奖", "美国", "EN", "https://x.com/ylecun", "世界模型派"],
    ["G.AI头部自媒体/KOL", "3", "Geoffrey Hinton", "@geoffreyhinton", "AI教父 诺奖", "加拿大", "EN", "https://x.com/geoffreyhinton", "AI风险派"],
    ["G.AI头部自媒体/KOL", "4", "Ilya Sutskever", "@ilyasut", "SSI创始人", "美国", "EN", "https://x.com/ilyasut", "极少发言但必转"],
    ["G.AI头部自媒体/KOL", "5", "Jeff Dean", "@JeffDean", "Google首席", "美国", "EN", "https://x.com/JeffDean", "Google技术风向"],
    ["G.AI头部自媒体/KOL", "6", "Demis Hassabis", "@demishassabis", "DeepMind CEO", "英国", "EN", "https://x.com/demishassabis", "诺奖得主"],
    ["G.AI头部自媒体/KOL", "7", "Jim Fan", "@DrJimFan", "NVIDIA总监", "美国", "EN", "https://x.com/DrJimFan", "具身智能+游戏AI"],
    ["G.AI头部自媒体/KOL", "8", "Andrew Ng", "@AndrewYNg", "DeepLearning.AI", "美国", "EN", "https://x.com/AndrewYNg", "AI教育布道者"],
    ["G.AI头部自媒体/KOL", "9", "Sam Altman", "@sama", "OpenAI CEO", "美国", "EN", "https://x.com/sama", "一手发布"],
    ["G.AI头部自媒体/KOL", "10", "Greg Brockman", "@gdb", "OpenAI总裁", "美国", "EN", "https://x.com/gdb", "技术发布"],
    ["G.AI头部自媒体/KOL", "11", "Elon Musk", "@elonmusk", "xAI", "美国", "EN", "https://x.com/elonmusk", "Grok相关必看"],
    ["G.AI头部自媒体/KOL", "12", "Aravind Srinivas", "@AravSrinivas", "Perplexity CEO", "美国", "EN", "https://x.com/AravSrinivas", "AI搜索一手"],
    ["G.AI头部自媒体/KOL", "13", "Clem Delangue", "@ClemDelangue", "HF CEO", "法国", "EN", "https://x.com/ClemDelangue", "开源社区"],
    ["G.AI头部自媒体/KOL", "14", "Rowan Cheung", "@rowancheung", "The Rundown 80万", "加拿大", "EN", "https://x.com/rowancheung", "AI日报流量王，必互动"],
    ["G.AI头部自媒体/KOL", "15", "Allie K. Miller", "@alliekmiller", "AI商业 100万", "美国", "EN", "https://x.com/alliekmiller", "AI商业评论"],
    ["G.AI头部自媒体/KOL", "16", "Ethan Mollick", "@emollick", "沃顿教授", "美国", "EN", "https://x.com/emollick", "AI与工作/教育"],
    ["G.AI头部自媒体/KOL", "17", "Deedy Das", "@deedydas", "创投数据", "美国", "EN", "https://x.com/deedydas", "数据可视化强"],
    ["G.AI头部自媒体/KOL", "18", "Hasan Toor", "@hasantoxr", "AI工具评测", "英国", "EN", "https://x.com/hasantoxr", "工具流"],
    ["G.AI头部自媒体/KOL", "19", "Santiago", "@svpino", "AI新闻", "美国", "EN", "https://x.com/svpino", "日更AI新闻"],
    ["G.AI头部自媒体/KOL", "20", "归藏", "@op7418", "华人AI自媒体", "中国/海外", "ZH", "https://x.com/op7418", "中文AI圈顶流"],
    ["G.AI头部自媒体/KOL", "21", "宝玉", "@dotey", "提示词/独立开发", "中国", "ZH", "https://x.com/dotey", "提示词专家"],
    ["G.AI头部自媒体/KOL", "22", "He Liang", "@tuture2", "AI产品", "中国/美国", "ZH/EN", "https://x.com/tuture2", "Lepton相关"],
]

ws = wb.active
ws.title = "AI_X关注清单"

# Title area
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "AI_X 关注清单 — 大模型 / AI工具 / SaaS / 中转站 / 媒体 / 评测 / KOL"
c.font = TITLE_FONT
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:I2")
c2 = ws["A2"]
c2.value = "共 120+ 官方X账号 · 带一键关注链接 · 更新至 2026-09-02 · 点击链接直达X关注 · 筛选列已开启"
c2.font = SUBTITLE_FONT
c2.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 16

# Header
headers = ["分类", "序号", "公司/账号", "X账号", "核心产品/模型", "地区", "语言", "一键关注链接", "备注/运营建议"]
col_widths = [22, 8, 22, 18, 26, 14, 10, 30, 46]
for idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
ws.row_dimensions[3].height = 22
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Data rows
start_row = 4
for r_idx, row in enumerate(data, start=start_row):
    cat = row[0]
    fill = CAT_FILL_MAP.get(cat, PatternFill(fill_type=None))
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name="Microsoft YaHei", size=9, color="1E293B")
        cell.alignment = Alignment(horizontal="left" if c_idx in (3,5,9) else "center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        # category column styling
        if c_idx == 1:
            cell.fill = fill
            cell.font = CAT_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # hyperlink for column 8
        if c_idx == 8 and val.startswith("http"):
            cell.hyperlink = val
            cell.font = Font(name="Microsoft YaHei", size=9, color="2563EB", underline="single")
            cell.value = val
        # handle column
        if c_idx == 4:
            cell.font = Font(name="Consolas", size=9, bold=True, color="0F172A")
    ws.row_dimensions[r_idx].height = 18
    # zebra
    if r_idx % 2 == 0:
        for c_idx in range(2, 10):
            cell = ws.cell(row=r_idx, column=c_idx)
            if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == "00000000":
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# Freeze and filter
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:I{start_row + len(data) -1}"
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A3
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_title_rows = "1:3"

# Add second sheet - guide
ws2 = wb.create_sheet("使用指南")
ws2.sheet_properties.pageSetUpPr.fitToPage = True
ws2.merge_cells("A1:B1")
ws2["A1"].value = "使用指南 - 如何高效用这份清单起号/引流"
ws2["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="0F172A")
guides = [
    ("1. 一键关注", "点击 H列 蓝色链接直接跳转X关注，建议按 分类筛选 分批关注，避免一次性关注过多被限流"),
    ("2. 优先级", "必关Top20：OpenAI / Anthropic / GoogleDeepMind / Meta AI / xAI / DeepSeek / HuggingFace / LM Arena / Artificial Analysis / Karpathy / Rowan Cheung 等"),
    ("3. 互动策略", "媒体/评测号适合转载+评论蹭流量；KOL适合深度回复+引用；大模型官方适合第一时间转载+解读"),
    ("4. 中转站", "OpenRouter/Portkey/Helicone 是做API分发的核心，SiliconFlow是国内合规替代"),
    ("5. 更新", "建议每周检查一次，X改名频繁，可用本表筛选后批量检查死链"),
    ("6. 筛选", "已开启自动筛选，点击表头可按 分类/地区/语言 筛选"),
]
for i, (k,v) in enumerate(guides, start=3):
    ws2.cell(row=i, column=1, value=k).font = Font(name="Microsoft YaHei", size=10, bold=True, color="1E293B")
    ws2.cell(row=i, column=2, value=v).font = Font(name="Microsoft YaHei", size=10, color="475569")
    ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    ws2.row_dimensions[i].height = 22
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 110
ws2.merge_cells("A10:B10")
ws2["A10"].value = "需要我再按 关注优先级 / 发帖频率 / 互动价值 打一个 Top20 必关精简版吗？"
ws2["A10"].font = Font(name="Microsoft YaHei", size=10, italic=True, color="64748B")

wb.save(filepath)
print(f"saved to {filepath} rows={len(data)}")
