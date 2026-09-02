import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

filepath = r"D:\web\StackHK13\AI_X_关注清单.xlsx"
wb = openpyxl.load_workbook(filepath)
ws = wb.active

# Define missing entries with metadata
missing_data = [
    # 评测机构/榜单
    ["H.评测资源补充(你的清单新增)", "", "LMSYS lmarena_ai", "@lmarena_ai", "Chatbot Arena 官方新号", "美国", "EN", "https://x.com/lmarena_ai", "原@lmsysorg已迁移至此，必关，实时榜单"],
    ["H.评测资源补充(你的清单新增)", "", "OpenCompass 司南", "@OpenCompassAI", "中文评测第一", "中国 上海", "ZH/EN", "https://x.com/OpenCompassAI", "中文MMLU/CEval核心榜"],
    ["H.评测资源补充(你的清单新增)", "", "HELM Stanford", "@stanfordnlp", "HELM评测", "美国", "EN", "https://x.com/stanfordnlp", "Stanford NLP官方"],
    ["H.评测资源补充(你的清单新增)", "", "SuperCLUE", "@superclue", "中文大模型评测", "中国", "ZH", "https://x.com/superclue", "中文评测"],
    ["H.评测资源补充(你的清单新增)", "", "CLUE社区", "@_clue_", "CLUE Benchmark", "中国", "ZH", "https://x.com/_clue_", "中文基准"],
    ["H.评测资源补充(你的清单新增)", "", "FlagEval 智源", "@FlagEval", "智源FlagEval", "中国 北京", "ZH", "https://x.com/FlagEval", "智源研究院榜单"],
    ["H.评测资源补充(你的清单新增)", "", "LiveBench 新号", "@livebenchAI", "LiveBench防污染榜", "美国", "EN", "https://x.com/livebenchAI", "与livebench_ai同榜，新号"],
    ["H.评测资源补充(你的清单新增)", "", "AlpacaEval", "@tatsu_lab", "AlpacaEval", "美国", "EN", "https://x.com/tatsu_lab", "Tatsu Lab维护"],
    ["H.评测资源补充(你的清单新增)", "", "C-Eval", "@CevalBench", "中文C-Eval", "中国", "ZH", "https://x.com/CevalBench", "中文知识评测"],
    ["H.评测资源补充(你的清单新增)", "", "CMMLU", "@haonanlcn", "CMMLU作者", "中国", "EN", "https://x.com/haonanlcn", "中文多任务"],
    ["H.评测资源补充(你的清单新增)", "", "MMLU-Pro", "@TIGERLabAI", "MMLU-Pro", "美国", "EN", "https://x.com/TIGERLabAI", "TIGER Lab"],
    ["H.评测资源补充(你的清单新增)", "", "SWE-Bench", "@swebench", "代码评测", "美国", "EN", "https://x.com/swebench", "Agent代码能力金标准"],
    ["H.评测资源补充(你的清单新增)", "", "EvalPlus HumanEval+", "@evalplusorg", "HumanEval+ ", "美国", "EN", "https://x.com/evalplusorg", "代码评测+"],
    ["H.评测资源补充(你的清单新增)", "", "LiveCodeBench", "@NamanM03", "LiveCodeBench", "美国", "EN", "https://x.com/NamanM03", "实时代码榜"],
    ["H.评测资源补充(你的清单新增)", "", "MathVista", "@mathvista_paper", "多模态数学", "美国", "EN", "https://x.com/mathvista_paper", "数学推理"],
    ["H.评测资源补充(你的清单新增)", "", "MMMU", "@MMMUBench", "多模态评测", "美国", "EN", "https://x.com/MMMUBench", "多学科多模态"],
    ["H.评测资源补充(你的清单新增)", "", "ARC-AGI", "@arcprize", "ARC Prize", "美国", "EN", "https://x.com/arcprize", "AGI抽象推理，奖金池大"],
    ["H.评测资源补充(你的清单新增)", "", "GAIA/SafetyBench", "@llm_bench", "SafetyBench", "中国", "EN", "https://x.com/llm_bench", "安全评测"],
    ["H.评测资源补充(你的清单新增)", "", "Needle-in-Haystack", "@gkamradt", "长上下文测试", "美国", "EN", "https://x.com/gkamradt", "Greg Kamradt"],
    # 个人权威专家
    ["H.评测资源补充(你的清单新增)", "", "Dan Hendrycks", "@DanHendrycks", "MMLU/MATH作者", "美国", "EN", "https://x.com/DanHendrycks", "Scale AI创始人，必关"],
    ["H.评测资源补充(你的清单新增)", "", "Mark Chen", "@markchen90", "HumanEval作者", "美国", "EN", "https://x.com/markchen90", "OpenAI"],
    ["H.评测资源补充(你的清单新增)", "", "Karl Cobbe", "@karlcobe", "GSM8K作者", "美国", "EN", "https://x.com/karlcobe", "OpenAI"],
    ["H.评测资源补充(你的清单新增)", "", "Lianmin Zheng", "@lianmin_zheng", "LMSYS创始人", "美国", "EN", "https://x.com/lianmin_zheng", "Arena核心"],
    ["H.评测资源补充(你的清单新增)", "", "Yueqi Li", "@yueqi0203", "MMLU-Pro维护", "美国", "EN", "https://x.com/yueqi0203", "MMLU-Pro"],
    ["H.评测资源补充(你的清单新增)", "", "Jason Wei", "@jxmnop", "思维链CoT作者", "美国", "EN", "https://x.com/jxmnop", "OpenAI，CoT鼻祖"],
    ["H.评测资源补充(你的清单新增)", "", "Harrison Chase", "@hwchase17", "LangChain创始人", "美国", "EN", "https://x.com/hwchase17", "Agent框架"],
    ["H.评测资源补充(你的清单新增)", "", "Aran Komatsuzaki", "@arankomatsuzaki", "论文速览", "日本/美国", "EN", "https://x.com/arankomatsuzaki", "每日论文"],
    ["H.评测资源补充(你的清单新增)", "", "Ahsen Khaliq", "@_akhaliq", "HF论文日报", "英国", "EN", "https://x.com/_akhaliq", "AK论文日报，必关"],
    ["H.评测资源补充(你的清单新增)", "", "Tri Dao", "@tri_dao", "FlashAttention作者", "美国", "EN", "https://x.com/tri_dao", "高效注意力"],
    ["H.评测资源补充(你的清单新增)", "", "Dario Amodei", "@DarioAmodei", "Anthropic CEO个人号", "美国", "EN", "https://x.com/DarioAmodei", "与@AnthropicAI互补"],
    ["H.评测资源补充(你的清单新增)", "", "Liam Li", "@liam_li", "MLPerf联合创始人", "美国", "EN", "https://x.com/liam_li", "评测基础设施"],
    ["H.评测资源补充(你的清单新增)", "", "Simon Willison", "@simonw", "独立AI评测博主", "英国", "EN", "https://x.com/simonw", "独立评测，工具流"],
    ["H.评测资源补充(你的清单新增)", "", "Maxime Labonne", "@maximelabonne", "LLM教程作者", "法国", "EN", "https://x.com/maximelabonne", "LLM教程"],
    ["H.评测资源补充(你的清单新增)", "", "Yannic Kilcher", "@ykilcher", "AI视频讲解", "瑞士", "EN", "https://x.com/ykilcher", "YouTube+论文解读"],
    ["H.评测资源补充(你的清单新增)", "", "Robert Scoble", "@Scobleizer", "AI行业评论", "美国", "EN", "https://x.com/Scobleizer", "行业观察"],
    ["H.评测资源补充(你的清单新增)", "", "Ben Thompson", "@benthompson", "Stratechery", "美国", "EN", "https://x.com/benthompson", "付费科技评论"],
    # 厂商/组织 增量
    ["H.评测资源补充(你的清单新增)", "", "智源 BAAI", "@BAAI_1988", "智源研究院", "中国 北京", "ZH", "https://x.com/BAAI_1988", "FlagEval主办"],
    ["H.评测资源补充(你的清单新增)", "", "阿里达摩院", "@AlibabaDAMO", "达摩院", "中国 杭州", "ZH", "https://x.com/AlibabaDAMO", "阿里研究"],
    ["H.评测资源补充(你的清单新增)", "", "百度文心", "@wenxin_baidu", "文心一言", "中国 北京", "ZH", "https://x.com/wenxin_baidu", "文心官方"],
    ["H.评测资源补充(你的清单新增)", "", "商汤 SenseTime", "@sensetime_group", "商汤", "中国 香港", "ZH/EN", "https://x.com/sensetime_group", "SenseNova"],
    ["H.评测资源补充(你的清单新增)", "", "月之暗面 MoonshotAI", "@MoonshotAI", "MoonshotAI官方", "中国 北京", "EN", "https://x.com/MoonshotAI", "与@Kimi_Moonshot区分，集团号"],
    ["H.评测资源补充(你的清单新增)", "", "面壁智能 ModelBest", "@modelbest_", "面壁智能", "中国 北京", "ZH", "https://x.com/modelbest_", "小钢炮模型"],
    # 研究机构/高校
    ["H.评测资源补充(你的清单新增)", "", "MIT CSAIL", "@MIT_CSAIL", "MIT CSAIL", "美国", "EN", "https://x.com/MIT_CSAIL", "MIT计算机与AI实验室"],
    ["H.评测资源补充(你的清单新增)", "", "Stanford AI Lab", "@StanfordAILab", "Stanford AI Lab", "美国", "EN", "https://x.com/StanfordAILab", "斯坦福AI"],
    ["H.评测资源补充(你的清单新增)", "", "Berkeley BAIR", "@BAIRlab", "BAIR", "美国", "EN", "https://x.com/BAIRlab", "伯克利AI"],
    ["H.评测资源补充(你的清单新增)", "", "CMU MLD", "@CMU_MLD", "CMU机器学习系", "美国", "EN", "https://x.com/CMU_MLD", "CMU"],
    ["H.评测资源补充(你的清单新增)", "", "Allen AI AI2", "@allenai_org", "Allen Institute", "美国", "EN", "https://x.com/allenai_org", "OLMo等"],
    ["H.评测资源补充(你的清单新增)", "", "MIT-IBM Watson", "@MITIBMWatsonAI", "MIT-IBM", "美国", "EN", "https://x.com/MITIBMWatsonAI", "联合实验室"],
    ["H.评测资源补充(你的清单新增)", "", "Microsoft Research", "@MSResearch", "微软研究院", "美国", "EN", "https://x.com/MSResearch", "MSR官方"],
    ["H.评测资源补充(你的清单新增)", "", "PKU 北大", "@PKU1898", "北京大学", "中国 北京", "ZH", "https://x.com/PKU1898", "北大官方"],
]

THIN_BORDER = Border(left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"), top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"))
CAT_FILL = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
CAT_FONT = Font(name="Microsoft YaHei", size=9, bold=True, color="9A3412")

start_row = ws.max_row + 1
for idx, row in enumerate(missing_data):
    r = start_row + idx
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c_idx, value=val)
        cell.font = Font(name="Microsoft YaHei", size=9, color="1E293B")
        cell.alignment = Alignment(horizontal="left" if c_idx in (3,5,9) else "center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if c_idx == 1:
            cell.fill = CAT_FILL
            cell.font = CAT_FONT
        if c_idx == 8 and val.startswith("http"):
            cell.hyperlink = val
            cell.font = Font(name="Microsoft YaHei", size=9, color="2563EB", underline="single")
        if c_idx == 4:
            cell.font = Font(name="Consolas", size=9, bold=True, color="0F172A")
    ws.row_dimensions[r].height = 18
    if r % 2 == 0:
        for c_idx in range(2,10):
            c = ws.cell(row=r, column=c_idx)
            if not c.fill.start_color.rgb or c.fill.start_color.rgb == "00000000":
                c.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# update filter
ws.auto_filter.ref = f"A3:I{ws.max_row}"
# update sheet title
ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(filepath)
print(f"appended {len(missing_data)} rows, total now {ws.max_row}")
